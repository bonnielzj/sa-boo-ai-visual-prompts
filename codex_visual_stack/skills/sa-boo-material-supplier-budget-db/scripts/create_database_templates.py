#!/usr/bin/env python3
"""Create SA&BOO material/supplier/budget database CSV and XLSX templates."""
from __future__ import annotations
import argparse, csv, pathlib
from datetime import date

SHEETS = {
    "Materials": [
        "Material ID","Category","Subcategory","Brand","Product Name","Model / SKU","Color / Finish","Size / Spec","Unit","Unit Price Low","Unit Price High","Currency","Default Supplier ID","Lead Time Days","MOQ","Sample Available","Certification","Maintenance","Sustainability Notes","Image URL / File","Source URL","Status","Risk Notes","Last Verified"
    ],
    "Suppliers": [
        "Supplier ID","Supplier Name","Category","Contact Person","Phone / WeChat","Email","City / Region","Website / Shop","Price Level","Quality Score","Reliability Score","Service Score","Lead Time Score","Documentation Score","Overall Score","Payment Terms","Warranty / After-sales","Cooperation Notes","Status","Last Contact"
    ],
    "Price_History": [
        "Price ID","Material ID","Supplier ID","Date","Unit Price","Currency","Tax Included","Freight Included","Install Included","Quote Valid Until","Source","Notes"
    ],
    "Project_Budget": [
        "Line ID","Project","Area / Room","Package","Item","Material ID","Supplier ID","Spec","Quantity","Unit","Unit Price Estimate","Quote Unit Price","Approved Unit Price","Tax Rate","Freight","Install Cost","Contingency %","Estimated Total","Quoted Total","Approved Total","Target Budget","Variance","Variance %","Approval Status","Procurement Status","Required Date","Risk Level","Notes"
    ],
    "Procurement_Tracker": [
        "Procurement ID","Project","Line ID","Supplier ID","PO / Order No","Order Date","Deposit Paid","Balance Paid","Expected Delivery","Actual Delivery","Install Date","Status","Issue","Next Action"
    ],
    "Approvals": [
        "Approval ID","Project","Line ID","Decision","Approved By","Date","Evidence","Notes"
    ],
    "Substitutions": [
        "Substitution ID","Project","Original Material ID","Replacement Material ID","Reason","Cost Impact","Schedule Impact Days","Design Impact","Approval Status","Notes"
    ],
    "Dashboard": ["Metric","Value","Notes"],
}

SAMPLES = {
    "Materials": [
        ["MAT-0001","Wall Finish","Art Paint","Example Brand","Warm Grey Limewash","LW-01","Warm grey matte","Interior wall finish","m2",180,320,"CNY","SUP-0001",14,"10m2","Yes","VOC report needed","Wipe gently, repair by area","Low VOC only if certificate provided","","https://example.com","active","color batch risk",str(date.today())],
        ["MAT-0002","Metal","Stainless Steel","Example Metal","Brushed Champagne Metal","SS-CH01","Champagne brushed","1.2mm sheet/custom trim","m",260,480,"CNY","SUP-0002",21,"custom","No","fire rating if required","avoid acidic cleaners","durable, recyclable claim requires proof","","https://example.com","watch","scratch/lead-time risk",str(date.today())],
    ],
    "Suppliers": [
        ["SUP-0001","Example Wall Finish Supplier","wall finish","Contact A","","","Shanghai","https://example.com","mid",4,4,4,3,3,"","50% deposit / 50% before delivery","1 year limited","Good sample response","trial",str(date.today())],
        ["SUP-0002","Example Metal Fabricator","metal","Contact B","","","Suzhou","https://example.com","high",4,3,4,3,3,"","custom quotation","case by case","Needs drawing confirmation","trial",str(date.today())],
    ],
    "Price_History": [
        ["PRICE-0001","MAT-0001","SUP-0001",str(date.today()),220,"CNY","No","No","No","","sample quote","verify before issue"],
    ],
    "Project_Budget": [
        ["PB-0001","Sample Project","Living Room","Wall Finish","Feature wall art paint","MAT-0001","SUP-0001","Warm grey matte",35,"m2",220,240,240,0.13,300,1200,0.08,"","","",10000,"","","pending","quoted","","medium","sample line"],
    ],
    "Procurement_Tracker": [],
    "Approvals": [],
    "Substitutions": [],
    "Dashboard": [
        ["Total Estimated","=SUM(Project_Budget!R:R)","formula in XLSX"],
        ["Total Quoted","=SUM(Project_Budget!S:S)","formula in XLSX"],
        ["Total Approved","=SUM(Project_Budget!T:T)","formula in XLSX"],
    ],
}

def write_csvs(out: pathlib.Path) -> None:
    for sheet, header in SHEETS.items():
        path = out / f"{sheet}.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(SAMPLES.get(sheet, []))

def write_xlsx(out: pathlib.Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
    except Exception as e:
        print(f"openpyxl unavailable, skipped XLSX: {e}")
        return
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="111111")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet, header in SHEETS.items():
        ws = wb.create_sheet(sheet)
        ws.append(header)
        for row in SAMPLES.get(sheet, []):
            ws.append(row)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        if len(ws["A"]) >= 2:
            end_col = get_column_letter(len(header))
            end_row = max(2, ws.max_row)
            tab = Table(displayName=f"tbl_{sheet}".replace(" ", "_"), ref=f"A1:{end_col}{end_row}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium1", showRowStripes=True, showColumnStripes=False)
            ws.add_table(tab)
        for idx, col in enumerate(header, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(max(len(col)+2, 12), 28)
    # formulas for sample project budget row
    ws = wb["Project_Budget"]
    # R Estimated Total, S Quoted Total, T Approved Total, V Variance, W Variance %
    for r in range(2, ws.max_row + 1):
        ws[f"R{r}"] = f"=I{r}*K{r}+O{r}+P{r}+(I{r}*K{r}+O{r}+P{r})*Q{r}+(I{r}*K{r}+O{r}+P{r})*(1+Q{r})*N{r}"
        ws[f"S{r}"] = f"=I{r}*L{r}+O{r}+P{r}+(I{r}*L{r}+O{r}+P{r})*Q{r}+(I{r}*L{r}+O{r}+P{r})*(1+Q{r})*N{r}"
        ws[f"T{r}"] = f"=I{r}*M{r}+O{r}+P{r}+(I{r}*M{r}+O{r}+P{r})*Q{r}+(I{r}*M{r}+O{r}+P{r})*(1+Q{r})*N{r}"
        ws[f"V{r}"] = f"=T{r}-U{r}"
        ws[f"W{r}"] = f"=IF(U{r}=0,0,V{r}/U{r})"
    # supplier score formula
    ws = wb["Suppliers"]
    for r in range(2, ws.max_row + 1):
        ws[f"O{r}"] = f"=J{r}*30%+K{r}*25%+L{r}*15%+M{r}*10%+N{r}*10%"
    wb.save(out / "SA&BOO_Material_Supplier_Budget_DB_Template.xlsx")

def main() -> None:
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--out", default="assets")
    args = p.parse_args()
    out = pathlib.Path(args.out).expanduser().resolve()
    out.mkdir(parents=True, exist_ok=True)
    write_csvs(out)
    write_xlsx(out)
    print(f"Generated SA&BOO database templates in {out}")

if __name__ == "__main__":
    main()
